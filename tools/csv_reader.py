import pandas as pd
import openpyxl

from .pig_factory import DongYingFactory, ParentError
from models.pig_model import PigModel
from .general import ask


class CsvReader:
    
    def __init__(self):
        self.df = None
        self.pigs = []
        self.review = None

class DongYingCsvReader(CsvReader):

    fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="DDDDDD")

    def __init__(self, path: str):

        super().__init__()
        print("Reading...")
        
        
        # df format:
        # [Index, Breed, ID, mark, Birthday, Sire, Dam, naif_id, Chinese_name, ., ., ., Gender,...]
        # to [Breed, ID, Birthday, Sire, Dam, naif_id, Chinese_name, Gender]
        self.df = pd.read_excel(
            path,
            usecols=[1,2,4,5,6,7,8,12],
            names=[
                'Breed',
                'ID',
                'Birthday',
                'Sire',
                'Dam',
                'naif_id',
                'Chinese_name',
                'Gender'
            ],
            dtype={
                'Breed':'object',
                'ID':'object',
                'Birthday':'object',
                'Sire':'object',
                'Dam':'object',
                'naif_id':'object',
                'Chinese_name':'object',
                'Gender':'object'
            }
        )
        # Clean empty rows.
        self.df.dropna(how = 'all', inplace = True)
        self.df_list = list(self.df.iterrows())
        print("Success")

    def create_pigs(self, ignore_parent: bool = False, in_farm: bool = True, nearest: bool = True, update: bool = False):
        '''
        Create pigs instance and generate list of pigs with incorrect format.
        
        If it is the first commit, please set `ignore_parent` as True since 
        parents are foriegn fields in the database.
        * param in_farm: `True` to only searching in_farm parent.
        * param nearest: `True` to automatically choose the pig with nearest birthday as the parent.
        * param ignore_parent: `False` to not arrange dam and sire of the pig.
        * param update: `True` to auto update when duplicate primary key is found.
        in the database in this commit.

        ## To Do
        * Check pigs with same id but close birthday.
        '''

        # Set output excel
        output = openpyxl.Workbook()
        sheet = output.create_sheet('基本資料')
        count = 0
        flags = [
            DongYingFactory.BREED_FLAG, 
            DongYingFactory.ID_FLAG, 
            DongYingFactory.BIRTHDAY_FLAG, 
            DongYingFactory.SIRE_FLAG,
            DongYingFactory.DAM_FLAG,
            DongYingFactory.NAIF_FLAG,
            DongYingFactory.GENDER_FLAG
        ]
        columns = ['a','b','c','d','e','f','h']
        parent_not_found = []
        model = PigModel()

        for index, pig in self.df_list:

            # Set values
            factory = DongYingFactory()
            not_nan = pig.notna()
            if not_nan[0]:
                factory.set_breed(str(pig.get('Breed')))
            if not_nan[1]:
                factory.set_id(str(pig.get('ID')))
            if not_nan[2]:
                factory.set_birthday(pig.get('Birthday').date())
            if not_nan[5]:
                factory.set_naif_id(str(pig.get('naif_id')))
            if not_nan[6]:
                factory.set_chinese_name(str(pig.get('Chinese_name')))
            if not_nan[7]:
                factory.set_gender(str(pig.get('Gender')))
            factory.set_farm()
        
            if not ignore_parent:
                try:
                    if not_nan[3]:
                        factory.set_parent(False, str(pig.get("Sire")), in_farm, nearest)
                    if not_nan[4]:
                        factory.set_parent(True, str(pig.get("Dam")), in_farm, nearest)
                except ParentError:
                    parent_not_found.append([factory, str(pig.get("Sire")), str(pig.get("Dam"))])
                    continue

            # If errors happen
            if factory.get_flag() > 0:
                count += 1
                data = pig.to_list()
                data.append(str(factory.error_messages))
                sheet.append(data)
                # Check the flag and highlight incorrect cells
                for i in range(len(flags)):
                    if factory.check_flag(flags[i]):
                        sheet[columns[i] + str(count)].fill = self.fill
            else:
                # If pig in pigs exists in the database, ask how to do.
                try:
                    if factory.pig.is_unique():
                        model.insert(factory.pig)
                except ValueError:
                    duplicate = model.find_pig(factory.pig)
                    if factory.pig == duplicate:
                        pass
                    else:
                        if not update:
                            if ask(
                                "遇到重複豬隻，請問如何處理？\n讀到的豬：\n{pig}\n已有的豬：\n{other}\nY:標記並略過 N：修改".format(
                                    pig=str(factory.pig),
                                    other=str(duplicate)
                                )
                            ):
                                # Incorrect data. Add to output excel.
                                data = pig.to_list()
                                data.append("與資料庫中的資料重複且不相符")
                                sheet.append(data)
                                count += 1
                            else:
                                model.update(factory.pig)
                        else:
                            #修改豬隻資料
                            model.update(factory.pig)
            
            factory = None
        
        while len(parent_not_found) > 0:
            left = []
            for pig in parent_not_found:
                try:
                    factory: DongYingFactory = pig[0]
                    if pig[1] != "nan":
                        factory.set_parent(False, pig[1], in_farm, nearest)
                    if pig[2] != "nan":
                        factory.set_parent(True, pig[2], in_farm, nearest)
                    if pig[1] == pig[2] and pig[1] == "nan":
                        left.append(pig)
                        continue
                    # insert
                    try:
                        if factory.pig.is_unique():
                            model.insert(factory.pig)
                    except ValueError:
                        duplicate = model.find_pig(factory.pig)
                        if factory.pig == duplicate:
                            pass
                        else:
                            if not update:
                                if ask(
                                    "遇到重複豬隻，請問如何處理？\n讀到的豬：\n{pig}\n已有的豬：\n{other}\nY:標記並略過 N：修改".format(
                                        pig=str(factory.pig),
                                        other=str(duplicate)
                                    )
                                ):
                                    # Incorrect data. Add to output excel.
                                    data = pig.to_list()
                                    data.append("與資料庫中的資料重複且不相符")
                                    sheet.append(data)
                                    count += 1
                                else:
                                    model.update(factory.pig)
                            else:
                                #修改豬隻資料
                                model.update(factory.pig)
                    factory = None
                except ParentError:
                    left.append(pig)
            if len(left) == len(parent_not_found):
                # Non of the pigs can be inserted
                break
            else:
                parent_not_found = left

        output.save('./test/output.xlsx')