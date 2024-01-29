from datetime import datetime
import openpyxl
import re

from data_structures.estrus import Estrus
from data_structures.estrus import PregnantStatus
from factory import ParentError
from factory.dong_ying_factory import DongYingPigFactory
from factory.dong_ying_factory import DongYingEstrusFactory
from factory.dong_ying_factory import DongYingMatingFactory
from factory.dong_ying_factory import DongYingFarrowingFactory
from models.pig_model import PigModel
from models.estrus_model import EstrusModel
from models.mating_model import MatingModel
from models.farrowing_model import FarrowingModel
from reader import ExcelReader
from general import ask
from general import type_check


class DongYingPigReader(ExcelReader):


    def __init__(self, path: str, output_file_name: str):

        type_check(path, "path", str)
        type_check(output_file_name, "output_file_name", str)

        # df format:
        # [Index, Breed, ID, mark, Birthday, Sire, Dam, reg_id, Chinese_name, ., ., ., Gender,...]
        # to [Breed, ID, Birthday, Sire, Dam, reg_id, Chinese_name, Gender]
        usecols=[1, 2, 3, 4, 5, 6, 7, 8, 12]
        names=[
            'Breed',
            'ID',
            "confusing_note",
            'Birthday',
            'Sire',
            'Dam',
            'reg_id',
            'Chinese_name',
            'Gender', 
        ]
        dtype={
            'Breed':'object',
            'ID':'object',
            "confusing_note": "object", 
            'Birthday':'object',
            'Sire':'object',
            'Dam':'object',
            'reg_id':'object',
            'Chinese_name':'object',
            'Gender':'object',
        }
        super().__init__(
            path=path, 
            use_columns=usecols, 
            names=names, 
            dtype=dtype,
            output_file_name=output_file_name,
            output_page_name="基本資料",
            flag_to_output={
                1: "a",
                2: "b",
                4: "d",
                8: "e",
                16: "f",
                32: "g",
                64: "h"
            }
        )

    def create_pigs(self, 
            ignore_parent: bool = False, 
            in_farm: bool = True, 
            nearest: bool = True, 
            update: bool = False
        ):
        '''
        Create pigs instance and generate list of pigs with incorrect format.

        If it is the first commit, please set `ignore_parent` as True since 
        parents are foriegn fields in the database.
        * param in_farm: `True` to only searching in_farm parent.
        * param nearest: `True` to automatically choose the pig with nearest birthday as the parent.
        * param ignore_parent: `False` to not arrange dam and sire of the pig.
        * param update: `True` to auto update when duplicate primary key is found.
        in the database in this commit.

        '''

        model = PigModel()

        # Control flow
        count_repeat = 0
        length = len(self.queue)

        while len(self.queue) != 0:

            # Control flow
            if count_repeat == length:
                if length == len(self.queue):
                    break
                else:
                    count_repeat = 0
                    length = len(self.queue)
            count_repeat += 1
            wait_for_parent = False

            # Set values
            factory = DongYingPigFactory()
            pig = self.queue.popleft()
            not_nan = pig.notna()
            if not_nan.get("Breed"):
                factory.set_breed(str(pig.get('Breed')))
            if not_nan.get("ID"):
                factory.set_id(str(pig.get('ID')))
            if not_nan.get("Birthday"):
                factory.set_birthday(pig.get('Birthday').date())
            if not_nan.get("reg_id"):
                factory.set_reg_id(str(pig.get('reg_id')))
            if not_nan.get("Chinese_name"):
                factory.set_chinese_name(str(pig.get('Chinese_name')))
            if not_nan.get("Gender"):
                factory.set_gender(str(pig.get('Gender')))
            factory.set_farm()

            if not ignore_parent:
                try:
                    if not_nan.get("Sire"):
                        factory.set_parent(False, str(pig.get("Sire")), in_farm, nearest)
                    if not_nan.get("Dam"):
                        factory.set_parent(True, str(pig.get("Dam")), in_farm, nearest)
                except ParentError:
                    self.queue.append(pig)
                    wait_for_parent = True

            # If any statement in the confusing note column, reject this pig.
            if not_nan.get("confusing_note"):
                factory._turn_on_flag(DongYingPigFactory.Flags.ID_FLAG.value)
                factory.error_messages.append("不允許有相近耳號")

            # If errors happen
            if factory.get_flag() != 0:
                self.insert_output(pig.to_list(), factory)
                if wait_for_parent:
                    self.queue.pop()
            else:
                # If pig in pigs exists in the database, ask how to do.
                try:
                    if wait_for_parent:
                        continue
                    if factory.pig.is_unique():
                        model.insert(factory.pig)
                except ValueError:
                    duplicate = model.find_pig(factory.pig)
                    if factory.pig == duplicate:
                        pass
                    else:
                        if not update:
                            if ask(
                                "遇到重複豬隻，請問如何處理？\n讀到的豬：\n{pig}\n已有的豬：\n{other}\nY:標記並略過 N：修改".format(
                                    pig=str(factory.pig),
                                    other=str(duplicate)
                                )
                            ):
                                # Incorrect data. Add to output excel.
                                factory.error_messages.append("與資料庫中的資料重複且不相符")
                                factory._turn_on_flag(factory.Flags.ID_FLAG)
                                self.insert_output(pig.to_list(), factory)
                            else:
                                model.update(factory.pig)
                        else:
                            #修改豬隻資料
                            model.update(factory.pig)

            factory = None

        super().end()


class DongYingEstrusAndMatingReader(ExcelReader):
    '''
    ## Description
    The reader to read estrus and mating data provided by Dong-Ying. 
    These data is included in a table, so I only implement one reader to 
    handle this job.

    ## Logic
    The excel only provides mating date. We have to come up a way to 
    determine following cases: 
    1. it is the second mating in a single estrus,
    2. the sow did not get pregnant in last estrus,
    3. it is in the next parity.

    We can distinguish these cases by calculating the time delta 
    between the two records. Here I define:
    1. -> delta = 0 ~ 21 days
    2. -> delta = 22 ~ 180 days
    3. -> delta > 180 days
    
    To calculate time delta, the reader will sort the records by mating_date 
    in ascending order, so the earlier record can be inserted into the 
    database first. Before insertion, the reader will check the last estrus 
    of the sow in the database to calculate the time delta. If the delta is 
    in range 1, then this record will only be inserted into the `mating` table.
    Delta in range 2 accompanying with different parity is seem as an error.

    ### Pregnant Status
    * If the `note` column mention "*流產*", then the status will be set to 
    abortion.
    * Else if 21th_day_test is `False`  or the note is "*未*上*", the status 
    will be set to no.
    * Else it will be set to unknown.

    ## Auto Repair
    Since there are plenty of emtpy columns in history data, so I come up with 
    some repair mechanisms, which are listed below.

    ### Parity
    * If the sow has historical estrus in the database, calculate time delta. 
    If delta > 180 days or pregnant = Yes, then parity = old parity + 1. 
    Else, the parity does not change.
    * If the sow has no historical data, then parity = 1.

    ### 21th_day_test
    * If the sow has another record in the excel and time delta <= 180 days, 
    it will be set to `False`. Else `True`.
    * This attribute has not included in the database. So this auto fill in 
    is not implemented. Here is just a note.


    '''

    def __init__(self, path: str):

        # df format:
        # [sow_id, parity, boar_id, mating_date, week_age, mating_time, times, character, 21th_day_test, note_date, note, ...]
        # to [sow_id, parity, boar_id, mating_date, mating_time, 21th_day_test, note]
        usecols = [0, 1, 2, 3, 5, 8, 11]
        names = [
            "sow_id", 
            "parity", 
            "boar_id", 
            "mating_date", 
            "mating_time", 
            "21th_day_test", 
            "note"
        ]
        dtype = {
            "sow_id": "object",
            # Parity may not be an integer in the excel.
            # Type should be checked later.
            "parity": "object", 
            "boar_id": "object",
            "mating_date": "object",
            "mating_time": "object",
            "21th_day_test": "object",
            "note": "object",
        }
        super().__init__(path, usecols, names, dtype)

        self.df.sort_values(by=["mating_date", "mating_time"], inplace=True)

    def create_estrus_and_mating(
            self,
            repair_parity: bool = False,
            repair_21th_day_test: bool = False,
            repair_pregnant_status: bool = False
    ):
        '''Please see the description of the class.'''

        # Set output excel
        output = openpyxl.Workbook()
        sheet = output.create_sheet("配種資料")
        count = 0

        estrus_model = EstrusModel()

        for index, record in self.df.iterrows():
            is_nan = record.isna()
            estrus_factory = DongYingEstrusFactory()

            # If sow_id or mating_date is empty, use a fake value so the factory know it is wrong.
            sow_id = "fake_id" if is_nan.iloc[0] else record.get("sow_id")
            mating_date = "1980-01-01" if is_nan.iloc[3] else str(record.get("mating_date").date())
            estrus_factory.set_sow(sow_id, mating_date)

            # If the mating_time column is empty, 00:00 is used.
            time = "00:00:00" if is_nan.iloc[4] else str(record.get("mating_time"))
            estrus_factory.set_estrus_datetime(mating_date, time)

            # Calculate time delta
            old_records = estrus_model.find_multiple(
                equal={"id": sow_id, "farm": "Dong-Ying"}, 
                smaller={"estrus_datetime": str(estrus_factory.estrus.get_estrus_datetime())},
                order_by="estrus_datetime DESC"
            )
            if len(old_records) == 0:
                delta = -1
            else:
                dt = datetime.strptime(" ".join([mating_date, time]), "%Y-%m-%d %H:%M")
                last: Estrus = old_records.pop()
                delta = (dt - last.get_estrus_datetime()).days

            # Check the type of parity
            if not isinstance(record.get("parity"), int):
                if repair_parity:
                    if delta == -1 or last.get_parity() is None:
                        estrus_factory.set_parity(1)
                    elif delta > 180 or last.get_pregnant() == PregnantStatus.YES:
                        estrus_factory.set_parity(last.get_parity() + 1)
                    else:
                        estrus_factory.set_parity(last.get_parity())
            else:
                estrus_factory.set_parity(int(record.get("parity")))

            # Check pregnant status
            estrus_factory.set_pregnant(PregnantStatus.UNKNOWN)
            note = "" if is_nan.iloc[6] else str(record.get("note"))
            if record.get("21th_day_test") == "x" \
                or (re.search(r"未.*上", note) is not None):
                estrus_factory.set_pregnant(PregnantStatus.NO)
            elif re.search(r"流產", note) is not None:
                estrus_factory.set_pregnant(PregnantStatus.ABORTION)

            # If the flag is on, throw this data to output file.
            if estrus_factory.get_flag() != 0:
                count += 1
                data = record.to_list()
                data.append(str(estrus_factory.error_messages))
                sheet.append(data)
                # Check the flag and highlight incorrect cells
                columns = {1:"A", 2:"D", 8:"B"}
                for flag in estrus_factory.Flags:
                    if estrus_factory.check_flag(flag.value):
                        sheet[columns[flag.value] + str(count)].fill = self.fill
                continue

            # Check time delta
            if delta == -1:
                # No previous record. Insert this one.
                estrus_model.insert(estrus_factory.estrus)
            elif delta <= 21:
                # It is the second mating. Do nothing.
                '''
                這裡應該要把接下來 mating 對應的 estrus 指向到 last 比較保險。
                ----------------------------NOT DONE---------------------------------
                '''
                pass
            elif 21 < delta <= 180:
                # Update last
                estrus_model.update_pregnant(last, PregnantStatus.NO)
                # Insert new
                estrus_model.insert(estrus_factory.estrus)
            else:
                # Update last
                estrus_model.update_pregnant(last, PregnantStatus.YES)
                # Insert new
                estrus_model.insert(estrus_factory.estrus)
                pass

            mating_factory = DongYingMatingFactory()
            mating_factory.set_estrus(estrus_factory.estrus)
            mating_factory.set_mating_datetime(mating_date, time)
            boar_id = "fake_id" if is_nan.iloc[2] else record.get("boar_id")
            mating_factory.set_boar(boar_id, mating_date)

            # If the flag is on, throw this data to output file.
            if mating_factory.get_flag() != 0:
                count += 1
                data = record.to_list()
                data.append(str(mating_factory.error_messages))
                sheet.append(data)
                # Check the flag and highlight incorrect cells
                columns = {1:"A", 2:"D", 4:"C"}
                for flag in mating_factory.Flags:
                    if mating_factory.check_flag(flag.value):
                        sheet[columns[flag.value] + str(count)].fill = self.fill
                continue

            MatingModel().insert(mating_factory.mating)

        output.save('./test/reader/dong_ying/output2.xlsx')


class DongYingFarrowingReader(ExcelReader):

    def __init__(self, path: str):

        cols = [0, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 20]
        names = [
            "id", 
            "farrowing_date", 
            "n_of_male", 
            "n_of_female", 
            "crushing", 
            "black", 
            "weak", 
            "malformation", 
            "dead", 
            "total_born", 
            "born_alive", 
            "born_dead", 
            "total_weight", 
            "note"
        ]
        dtype = {key: value for key, value in zip(names, ["object"] * len(names))}
        super().__init__(path, cols, names, dtype)

    def create_farrowings(self):
        '''
        total_born, born_alive and born_dead will be checked.
        '''

        # Set output excel
        output = openpyxl.Workbook()
        sheet = output.create_sheet("配種資料")
        count = 0

        model = FarrowingModel()

        for index, record in self.df.iterrows():

            is_nan = record.isna()
            factory = DongYingFarrowingFactory()

            id = "fake_id" if is_nan.iloc[0] else record.get("id")
            farrowing_date = "1980-01-01" if is_nan.iloc[1] else record.get("farrowing_date").date()
            factory.set_estrus(id, farrowing_date)

            factory.set_farrowing_date(farrowing_date)

            def set_numeric(func, name, i, chinese, flag):
                if not is_nan.iloc[i]:
                    if isinstance(record.get(name), int):
                        func(int(record.get(name)))
                    else:
                        factory._turn_on_flag(flag)
                        factory.error_messages.append("{chinese}數字格式錯誤".format(chinese=chinese))

            set_numeric(factory.set_n_of_male, "n_of_male", 2, "公小豬", factory.Flags.N_OF_MALE_FLAG.value)
            set_numeric(factory.set_n_of_female, "n_of_female", 3, "母小豬", factory.Flags.N_OF_FEMALE_FLAG.value)
            set_numeric(factory.set_crushing, "crushing", 4, "壓死", factory.Flags.CRUSHING_FLAG.value)
            set_numeric(factory.set_black, "black", 5, "黑胎", factory.Flags.BLACK_FLAG.value)
            set_numeric(factory.set_weak, "weak", 6, "虛弱死", factory.Flags.WEAK_FLAG.value)
            set_numeric(factory.set_malformation, "malformation", 7, "畸形", factory.Flags.MALFORMATION_FLAG.value)
            set_numeric(factory.set_dead, "dead", 8, "死胎", factory.Flags.DEAD_FLAG.value)

            born_alive = 0 if is_nan.iloc[10] else int(record.get("born_alive"))
            born_dead = 0 if is_nan.iloc[11] else int(record.get("born_dead"))
            if born_dead != factory.farrowing.get_born_dead():
                factory._turn_on_flag(factory.Flags.CRUSHING_FLAG.value)
                factory._turn_on_flag(factory.Flags.BLACK_FLAG.value)
                factory._turn_on_flag(factory.Flags.WEAK_FLAG.value)
                factory._turn_on_flag(factory.Flags.MALFORMATION_FLAG.value)
                factory._turn_on_flag(factory.Flags.DEAD_FLAG.value)
                factory.error_messages.append("總死仔數與死亡數不相符")
            if born_alive != factory.farrowing.get_born_alive():
                factory._turn_on_flag(factory.Flags.N_OF_MALE_FLAG.value)
                factory._turn_on_flag(factory.Flags.N_OF_FEMALE_FLAG.value)
                factory.error_messages.append("活仔數與出生公母小豬數不相符")

            set_numeric(factory.set_total_weight, "total_weight", 12, "出生窩重", factory.Flags.TOTAL_WEIGHT_FLAG)
            if not is_nan.iloc[13]:
                factory.set_note(record.get("note"))

            # If the flag is on, throw this data to output file.
            if factory.get_flag() != 0:
                count += 1
                data = record.to_list()
                data.append(str(factory.error_messages))
                sheet.append(data)
                # Check the flag and highlight incorrect cells
                columns = {
                    1: "A", 
                    2: "B", 
                    4: "E", 
                    8: "F", 
                    16: "G", 
                    32: "H", 
                    64: "I", 
                    128: "C",
                    256: "D", 
                    512: "M"
                }
                for flag in factory.Flags:
                    if factory.check_flag(flag.value):
                        sheet[columns[flag.value] + str(count)].fill = self.fill
                continue

            model.insert(factory.farrowing)
            EstrusModel().update_pregnant(factory.farrowing.get_estrus(), PregnantStatus.YES)

        output.save('./test/reader/dong_ying/output3.xlsx')