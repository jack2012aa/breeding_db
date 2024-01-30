import pandas as pd
import openpyxl
from collections import deque

from factory import Factory
from general import type_check

class ExcelReader:

    fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="DDDDDD")

    def __init__(
            self, 
            path: str, 
            use_columns: list, 
            names: list, 
            dtype: dict, 
            output_file_name: str,
            output_page_name: str, 
            flag_to_output: dict, 
            not_null: bool
    ):
        '''
        * param path: the path of the excel file.
        * param usecolumns: which columns are to read in the excel.
        * param names: the name of these columns used in the dataframe.
        * param dtype: dtype of these columns.
        * param output_file_name: the file name of the excel noting incorrect data.
        * param output_page_name: the page name of the excel noting incorrect data.
        * param flag_to_output: a dictionary maps flag value to output column.
        * param not_null: whether the reader allows null value.
        '''

        type_check(path, "path", str)
        type_check(output_file_name, "output_file_name", str)
        type_check(output_page_name, "output_page_name", str)
        type_check(not_null, "not_null", bool)

        print("Reading...")
        self.df = pd.read_excel(
            path,
            usecols=use_columns,
            names=names,
            dtype=dtype
        )
        # Clean empty rows.
        self.df.dropna(how="all", inplace=True)
        self.queue = deque()
        for index, row in self.df.iterrows():
            self.queue.append(row)
        print("Success")

        # Set output excel.
        self.__output = openpyxl.Workbook()
        self.__sheet = self.__output.create_sheet(output_page_name)
        names.append("註釋")
        self.__sheet.append(names)
        self.__count = 2
        self.__output_columns = flag_to_output
        self.__output_file_name = output_file_name

        # Other attributes.
        self._factory: Factory = None
        self._not_null = not_null
        self._record: pd.Series = None
        self._not_nan: pd.Series = None

    def _set_output_columns(self, flag_to_output):
        self.__output_columns = flag_to_output

    def insert_output(self):
        '''Insert error messages in the factory to the sheet and return current count.'''

        data = self._record.to_list()
        data.append(str(self._factory.error_messages))
        self.__sheet.append(data)
        # Check the flag and highlight incorrect cells
        for flag in self._factory.Flags:
            if self._factory.check_flag(flag.value):
                self.__sheet[self.__output_columns[flag.value] + str(self.__count)].fill = self.fill
        
        self.__count += 1

    def end(self):

        self.__output.save("{name}.xlsx".format(name=self.__output_file_name))

    def _check_null(self, field: str, flag: int, default = None):
        '''
        Check whether the `field` is nan in the `record` Series. 
        If `not_null`, this function will turn on `flag`. 
        If the field is nan, return default.
        * param field: name of the field in `record` Series.
        * param flag: value of the flag.
        * param default: default value when the field is empty.
        '''

        if self._not_nan.get(field):
            return self._record.get(field)
        elif self._not_null:
            self._factory._turn_on_flag(flag)
            self._factory.error_messages.append("{field} 不能為空值".format(field=field))
        return default